#!/usr/bin/env python3
"""Build an Excel workbook from a curated gene-list CSV."""

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def style_sheet(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for idx, width in {
        1: 18,
        2: 28,
        3: 16,
        4: 22,
        5: 42,
        6: 18,
        7: 18,
        8: 72,
        9: 16,
        10: 28,
    }.items():
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Curated CSV input")
    parser.add_argument("--output", required=True, help="Workbook output path")
    parser.add_argument("--group-column", default="Inclusion tier", help="Column used to split sheets")
    args = parser.parse_args()

    with open(args.input, "r", newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)
        headers = reader.fieldnames or []

    grouped = defaultdict(list)
    for row in rows:
        group = (row.get(args.group_column) or "All").strip() or "All"
        grouped[group].append(row)

    workbook = Workbook()
    first = True
    for group_name, group_rows in grouped.items():
        ws = workbook.active if first else workbook.create_sheet()
        first = False
        ws.title = group_name[:31]
        ws.append(headers)
        for row in group_rows:
            ws.append([row.get(header, "") for header in headers])
        style_sheet(ws)

    notes = workbook.create_sheet("Notes")
    notes.append(["Input", str(Path(args.input).resolve())])
    notes.append(["Group column", args.group_column])
    notes.column_dimensions["A"].width = 18
    notes.column_dimensions["B"].width = 100
    for row in notes.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
